"""
Bulk import services: field auto-detection and row-by-row processing.
"""

from __future__ import annotations

import csv
from datetime import date
from decimal import Decimal, InvalidOperation
import io
from typing import IO, Any

import structlog

from .constants import (
    FIELD_ALIASES,
    IMPORT_FIELD_DEFINITIONS,
    FileType,
    ImportLogStatus,
    ImportStatus,
    ImportType,
)


logger = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# File parsing helpers
# ---------------------------------------------------------------------------


def _read_headers_csv(file_obj: IO[bytes]) -> list[str]:
    """Return the first-row headers from a CSV file."""
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))
    try:
        return next(reader)
    except StopIteration:
        return []


def _read_headers_xlsx(file_obj: IO[bytes]) -> list[str]:
    """Return first-row headers from an xlsx file using openpyxl."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        logger.warning("openpyxl not installed; xlsx detection unavailable")
        return []

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell) if cell is not None else "" for cell in row]
        break
    wb.close()
    return headers


def read_file_headers(file_obj: IO[bytes], file_type: str) -> list[str]:
    """Read column headers from a CSV or Excel file without loading the entire file."""
    if file_type == FileType.XLSX:
        return _read_headers_xlsx(file_obj)
    return _read_headers_csv(file_obj)


def _iter_rows_csv(file_obj: IO[bytes], mapping: dict[str, str]) -> list[dict[str, Any]]:
    """Yield dicts of {field_name: value} for each data row in a CSV."""
    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        mapped: dict[str, Any] = {}
        for col, field in mapping.items():
            if field and col in row:
                mapped[field] = row[col]
        rows.append(mapped)
    return rows


def _iter_rows_xlsx(file_obj: IO[bytes], mapping: dict[str, str]) -> list[dict[str, Any]]:
    """Yield dicts of {field_name: value} for each data row in an xlsx file."""
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        logger.warning("openpyxl not installed; xlsx reading unavailable")
        return []

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(cell) if cell is not None else "" for cell in row]
            continue
        row_dict = dict(zip(headers, row, strict=False))
        mapped: dict[str, Any] = {}
        for col, field in mapping.items():
            if field and col in row_dict:
                mapped[field] = row_dict[col]
        rows.append(mapped)
    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Auto-detect field mapping
# ---------------------------------------------------------------------------


def detect_field_mapping(headers: list[str], import_type: str) -> dict[str, str | None]:
    """
    Given a list of column headers and an import type, return a dict mapping
    each header to the best-matching model field name (or None if no match found).

    Detection is done via normalised string comparison against FIELD_ALIASES.
    """
    field_defs = IMPORT_FIELD_DEFINITIONS.get(import_type, {})
    # Build a reverse lookup: normalised alias → field_name
    alias_to_field: dict[str, str] = {}
    for field_name in field_defs:
        aliases = FIELD_ALIASES.get(field_name, [field_name])
        for alias in aliases:
            alias_to_field[alias.lower().strip()] = field_name

    result: dict[str, str | None] = {}
    for header in headers:
        normalised = header.lower().strip().replace("-", "_").replace(" ", "_")
        # Try exact-normalised match first
        match = alias_to_field.get(normalised)
        if not match:
            # Try space-replaced version
            normalised_spaced = header.lower().strip()
            match = alias_to_field.get(normalised_spaced)
        result[header] = match
    return result


# ---------------------------------------------------------------------------
# Row validators / coercers
# ---------------------------------------------------------------------------


def _coerce_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "ya", "y")
    return False


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if not value:
        return None
    from datetime import datetime  # noqa: PLC0415

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()  # noqa: DTZ007
        except ValueError:
            continue
    return None


def _coerce_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except InvalidOperation:
        return None


# ---------------------------------------------------------------------------
# Row-level import handlers per ImportType
# ---------------------------------------------------------------------------


def _import_employee_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.employee.models import Employee  # noqa: PLC0415

    employee_id = str(data.get("employee_id", "")).strip()
    if not employee_id:
        return False, "employee_id is required"

    dob = _coerce_date(data.get("date_of_birth"))
    if dob is None:
        return False, "date_of_birth is required and must be a valid date"

    defaults: dict[str, Any] = {
        "first_name": str(data.get("first_name", "")).strip(),
        "last_name": str(data.get("last_name", "")).strip(),
        "date_of_birth": dob,
        "gender": str(data.get("gender", "other")).strip().lower(),
        "nationality": str(data.get("nationality", "")).strip(),
        "national_id": str(data.get("national_id", "")).strip(),
        "phone": str(data.get("phone", "")).strip(),
        "personal_email": str(data.get("personal_email", "")).strip(),
        "address_line_1": str(data.get("address_line_1", "")).strip(),
        "address_line_2": str(data.get("address_line_2", "")).strip(),
        "city": str(data.get("city", "")).strip(),
        "state_province": str(data.get("state_province", "")).strip(),
        "postal_code": str(data.get("postal_code", "")).strip(),
        "country": str(data.get("country", "")).strip(),
    }

    if not defaults["first_name"] or not defaults["last_name"]:
        return False, "first_name and last_name are required"

    Employee.objects.update_or_create(employee_id=employee_id, defaults=defaults)
    return True, ""


def _import_department_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.organization.models import Department  # noqa: PLC0415

    name = str(data.get("name", "")).strip()
    code = str(data.get("code", "")).strip()
    if not name or not code:
        return False, "name and code are required"

    parent = None
    parent_code = str(data.get("parent_code", "")).strip()
    if parent_code:
        try:
            parent = Department.objects.get(code=parent_code)
        except Department.DoesNotExist:
            return False, f"parent department with code '{parent_code}' not found"

    Department.objects.update_or_create(
        code=code,
        defaults={
            "name": name,
            "description": str(data.get("description", "")).strip(),
            "parent": parent,
        },
    )
    return True, ""


def _import_job_title_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.organization.models import Department, JobTitle  # noqa: PLC0415

    name = str(data.get("name", "")).strip()
    code = str(data.get("code", "")).strip()
    if not name or not code:
        return False, "name and code are required"

    department = None
    dept_code = str(data.get("department_code", "")).strip()
    if dept_code:
        try:
            department = Department.objects.get(code=dept_code)
        except Department.DoesNotExist:
            return False, f"department with code '{dept_code}' not found"

    JobTitle.objects.update_or_create(
        code=code,
        defaults={
            "name": name,
            "description": str(data.get("description", "")).strip(),
            "department": department,
        },
    )
    return True, ""


def _import_employment_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.employee.models import Employee  # noqa: PLC0415
    from applications.employment.models import Employment  # noqa: PLC0415
    from applications.organization.models import Department, JobTitle  # noqa: PLC0415

    employee_id = str(data.get("employee_id", "")).strip()
    try:
        employee = Employee.objects.get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return False, f"employee with ID '{employee_id}' not found"

    dept_code = str(data.get("department_code", "")).strip()
    try:
        department = Department.objects.get(code=dept_code)
    except Department.DoesNotExist:
        return False, f"department with code '{dept_code}' not found"

    title_code = str(data.get("job_title_code", "")).strip()
    try:
        job_title = JobTitle.objects.get(code=title_code)
    except JobTitle.DoesNotExist:
        return False, f"job title with code '{title_code}' not found"

    hire_date = _coerce_date(data.get("hire_date"))
    if hire_date is None:
        return False, "hire_date is required and must be a valid date"

    end_date = _coerce_date(data.get("end_date"))

    Employment.objects.update_or_create(
        employee=employee,
        defaults={
            "department": department,
            "job_title": job_title,
            "work_location": str(data.get("work_location", "onsite")).strip().lower(),
            "status": str(data.get("status", "active")).strip().lower(),
            "hire_date": hire_date,
            "end_date": end_date,
        },
    )
    return True, ""


def _import_contract_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.employee.models import Employee  # noqa: PLC0415
    from applications.employment.models import Contract, Employment  # noqa: PLC0415

    employee_id = str(data.get("employee_id", "")).strip()
    try:
        employee = Employee.objects.get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return False, f"employee with ID '{employee_id}' not found"

    try:
        employment = Employment.objects.get(employee=employee)
    except Employment.DoesNotExist:
        return False, f"employment record for employee '{employee_id}' not found"

    start_date = _coerce_date(data.get("start_date"))
    if start_date is None:
        return False, "start_date is required and must be a valid date"

    end_date = _coerce_date(data.get("end_date"))

    Contract.objects.create(
        employment=employment,
        contract_type=str(data.get("contract_type", "permanent")).strip().lower(),
        start_date=start_date,
        end_date=end_date,
        terms=str(data.get("terms", "")).strip(),
    )
    return True, ""


def _import_salary_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.employee.models import Employee  # noqa: PLC0415
    from applications.employment.models import Employment, Salary  # noqa: PLC0415

    employee_id = str(data.get("employee_id", "")).strip()
    try:
        employee = Employee.objects.get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return False, f"employee with ID '{employee_id}' not found"

    try:
        employment = Employment.objects.get(employee=employee)
    except Employment.DoesNotExist:
        return False, f"employment record for employee '{employee_id}' not found"

    amount = _coerce_decimal(data.get("amount"))
    if amount is None:
        return False, "amount is required and must be a valid number"

    effective_date = _coerce_date(data.get("effective_date"))
    if effective_date is None:
        return False, "effective_date is required and must be a valid date"

    Salary.objects.create(
        employment=employment,
        pay_grade=str(data.get("pay_grade", "")).strip(),
        amount=amount,
        currency=str(data.get("currency", "IDR")).strip().upper(),
        payment_frequency=str(data.get("payment_frequency", "monthly")).strip().lower(),
        effective_date=effective_date,
        end_date=_coerce_date(data.get("end_date")),
        notes=str(data.get("notes", "")).strip(),
    )
    return True, ""


def _import_emergency_contact_row(data: dict[str, Any]) -> tuple[bool, str]:
    from applications.employee.models import EmergencyContact, Employee  # noqa: PLC0415

    employee_id = str(data.get("employee_id", "")).strip()
    try:
        employee = Employee.objects.get(employee_id=employee_id)
    except Employee.DoesNotExist:
        return False, f"employee with ID '{employee_id}' not found"

    name = str(data.get("name", "")).strip()
    relationship = str(data.get("relationship", "other")).strip().lower()
    phone = str(data.get("phone", "")).strip()

    if not name or not phone:
        return False, "name and phone are required"

    EmergencyContact.objects.create(
        employee=employee,
        name=name,
        relationship=relationship,
        phone=phone,
        email=str(data.get("email", "")).strip(),
        address=str(data.get("address", "")).strip(),
        is_primary=_coerce_bool(data.get("is_primary", False)),
    )
    return True, ""


_ROW_HANDLERS = {
    ImportType.EMPLOYEE: _import_employee_row,
    ImportType.EMERGENCY_CONTACT: _import_emergency_contact_row,
    ImportType.EMPLOYMENT: _import_employment_row,
    ImportType.CONTRACT: _import_contract_row,
    ImportType.SALARY: _import_salary_row,
    ImportType.DEPARTMENT: _import_department_row,
    ImportType.JOB_TITLE: _import_job_title_row,
}


# ---------------------------------------------------------------------------
# Main processing entry point
# ---------------------------------------------------------------------------


def _determine_final_status(success_rows: int, failed_rows: int) -> ImportStatus:
    if failed_rows == 0:
        return ImportStatus.COMPLETED
    if success_rows == 0:
        return ImportStatus.FAILED
    return ImportStatus.PARTIAL


def _read_file_rows(job: Any) -> list[dict[str, Any]]:
    job.file.open("rb")
    try:
        if job.file_type == FileType.XLSX:
            return _iter_rows_xlsx(job.file, job.field_mapping)
        return _iter_rows_csv(job.file, job.field_mapping)
    finally:
        job.file.close()


def process_import_job(job_id: str) -> None:
    """
    Read the uploaded file, apply the saved field mapping, and insert/update
    records row by row. Updates the ImportJob status and counters in-place.
    """
    from .models import ImportJob, ImportLog  # noqa: PLC0415

    try:
        job = ImportJob.objects.get(pk=job_id)
    except ImportJob.DoesNotExist:
        logger.error("ImportJob not found", job_id=job_id)
        return

    if not job.field_mapping:
        job.status = ImportStatus.FAILED
        job.error_summary = "Field mapping is not configured."
        job.save(update_fields=["status", "error_summary"])
        return

    handler = _ROW_HANDLERS.get(job.import_type)
    if handler is None:
        job.status = ImportStatus.FAILED
        job.error_summary = f"No handler found for import type '{job.import_type}'."
        job.save(update_fields=["status", "error_summary"])
        return

    job.status = ImportStatus.PROCESSING
    job.save(update_fields=["status"])

    rows = _read_file_rows(job)
    job.total_rows = len(rows)
    job.processed_rows = 0
    job.success_rows = 0
    job.failed_rows = 0
    job.save(update_fields=["total_rows", "processed_rows", "success_rows", "failed_rows"])

    logs: list[ImportLog] = []
    for row_num, row_data in enumerate(rows, start=2):
        success, error_msg = _process_single_row(handler, row_data, row_num, str(job_id))
        log_status = ImportLogStatus.SUCCESS if success else ImportLogStatus.ERROR
        logs.append(ImportLog(job=job, row_number=row_num, status=log_status, raw_data=row_data, error_message=error_msg))
        job.processed_rows += 1
        if success:
            job.success_rows += 1
        else:
            job.failed_rows += 1

    ImportLog.objects.bulk_create(logs)
    job.status = _determine_final_status(job.success_rows, job.failed_rows)
    job.save(update_fields=["processed_rows", "success_rows", "failed_rows", "status"])
    logger.info(
        "Import job finished",
        job_id=str(job_id),
        total=job.total_rows,
        success=job.success_rows,
        failed=job.failed_rows,
        status=job.status,
    )


def _process_single_row(handler: Any, row_data: dict[str, Any], row_num: int, job_id: str) -> tuple[bool, str]:
    try:
        return handler(row_data)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Unexpected error on row", row=row_num, job_id=job_id, error=str(exc))
        return False, str(exc)
